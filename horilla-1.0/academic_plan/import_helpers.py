"""
academic_plan.import_helpers
============================
Parser de Excel de Plan de Estudios.

Soporta los formatos reales que entrega Elysa:

Formato A — "ficha MBA" con hoja `Hoja1` (título + header en fila 5) y `sheet1`
  con headers en fila 1:
    Campus, Prog Acad, Descripción, Plan Acad, Lista Cso, Descripción, Descr,
    Semestre, Catálogo, ID Curso, Asignatura, Créditos, Hrs Semana, Hrs Semestre

Formato B — hoja única `sheet1`, fila 1 = título "Plan de Estudio", fila 2 headers:
    Institución, Campus, Grado, Prog Acad, Descripción, Plan Acad, Lista Cso,
    Descr, ID Curso, Nom Largo, Máx Uni, Hrs Cso, Catálogo, Estado, Cso Equiv

  En este formato el semestre está embebido en `Descr`
  (ej. "MMEPI20212 I Semestre") — se extrae por regex.
"""
from __future__ import annotations
import io
import re
import datetime as _dt
from openpyxl import load_workbook


def _norm(v):
    if v is None:
        return ""
    return str(v).strip()


def _find_header_row(ws, max_search=15):
    """Encuentra la fila con columnas de plan. Acepta tanto el formato A
    (asignatura/creditos) como el B (nom largo / máx uni)."""
    KEYS_ANY = (
        "semestre", "catálogo", "catalogo", "asignatura", "creditos", "créditos",
        "nom largo", "máx uni", "max uni", "hrs cso", "prog acad", "plan acad",
        "id curso",
    )
    for i in range(1, max_search + 1):
        row = [_norm(ws.cell(row=i, column=c).value).lower()
               for c in range(1, ws.max_column + 1)]
        joined = " | ".join(row)
        if sum(1 for k in KEYS_ANY if k in joined) >= 3:
            return i
    return 1


def _pick_data_sheet(wb):
    """Hoja con más celdas — los Excel reales tienen 'sheet1' con más data."""
    best = wb.active
    best_size = (best.max_row or 0) * (best.max_column or 0)
    for ws in wb.worksheets:
        size = (ws.max_row or 0) * (ws.max_column or 0)
        if size > best_size:
            best = ws
            best_size = size
    return best


def _detect_program_code(wb, fallback: str = "") -> tuple[str, str]:
    """Lee de Hoja1 (formato A) el código y nombre del programa."""
    if "Hoja1" not in wb.sheetnames:
        return fallback, ""
    ws = wb["Hoja1"]
    nombre = code = ""
    for r in range(1, min(8, ws.max_row + 1)):
        for c in range(1, min(8, ws.max_column + 1)):
            v = _norm(ws.cell(row=r, column=c).value)
            if v in ("Programa Académico", "Programa Academico"):
                try:
                    nombre = _norm(ws.cell(row=r + 1, column=c).value)
                except Exception:
                    pass
            elif v in ("Plan Académico", "Plan Academico", "Código", "Codigo"):
                try:
                    code = _norm(ws.cell(row=r + 1, column=c).value)
                except Exception:
                    pass
    if not code:
        for r in range(1, 8):
            for c in range(1, ws.max_column + 1):
                v = _norm(ws.cell(row=r, column=c).value)
                if 7 <= len(v) <= 12 and any(ch.isdigit() for ch in v) and any(ch.isupper() for ch in v):
                    code = v
                    break
    short = (code[:5] or fallback or "PROG").upper()
    return short, nombre


# Mapeo canónico de encabezados (todo en minúsculas).
COLMAP = {
    # Formato A
    "campus": "campus",
    "prog acad": "programa_codigo",
    "código programa": "programa_codigo",
    "codigo programa": "programa_codigo",
    "plan acad": "cohorte",
    "plan académico": "cohorte",
    "semestre": "semestre",
    "catálogo": "catalogo",
    "catalogo": "catalogo",
    "asignatura": "asignatura",
    "creditos": "creditos",
    "créditos": "creditos",
    "hrs semana": "hrs_semanal",
    "hrs semanal": "hrs_semanal",
    "hrs semestre": "hrs_semestre",
    "atributo curso": "atrib_curso",
    "atrib curso": "atrib_curso",
    "modalidad": "modalidad",
    "grado": "grado",
    # Formato B
    "nom largo": "asignatura",
    "máx uni": "creditos",
    "max uni": "creditos",
    "hrs cso": "hrs_semestre",
    "descr": "_descr",          # auxiliar — para extraer el semestre
    "institución": "_institucion",
    "institucion": "_institucion",
    "id curso": "_id_curso",
    "estado": "_estado",
}


_SEMESTRE_RE = re.compile(
    r"\b(I{1,4}|V|VI{0,3}|IX|X|XI{0,2}|[0-9]{1,2})\s*-?\s*[Ss]emestre\b"
)


def _parse_semestre_desde_descr(desc: str) -> str:
    """Ej.: 'MMEPI20212 I Semestre' → 'I'; 'ESGSA20251 II Semestre' → 'II'."""
    if not desc:
        return ""
    m = _SEMESTRE_RE.search(desc)
    if m:
        return m.group(1)
    # fallback: último token antes de 'Semestre'
    parts = desc.split()
    for i, p in enumerate(parts):
        if p.lower().startswith("semestre") and i > 0:
            return parts[i - 1]
    return ""


def parse_plan(file_bytes: bytes, filename: str) -> list[dict]:
    """Devuelve lista de dicts con los campos del modelo PlanEstudio."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _pick_data_sheet(wb)

    header_row = _find_header_row(ws)
    headers_raw = [_norm(ws.cell(row=header_row, column=c).value)
                   for c in range(1, ws.max_column + 1)]
    headers = [h.lower() for h in headers_raw]

    # Mapeo idx → field (primer match gana)
    field_by_idx: dict[int, str] = {}
    for i, h in enumerate(headers):
        field = COLMAP.get(h)
        if field and i not in field_by_idx:
            field_by_idx[i] = field

    code_default, _nombre = _detect_program_code(
        wb, fallback=filename.split('.')[0][:5].upper()
    )

    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in row):
            continue

        kw: dict = {}
        aux: dict = {}  # campos auxiliares (_descr, _estado, etc.)

        for i, field in field_by_idx.items():
            if i >= len(row):
                continue
            v = row[i]
            if v in (None, ""):
                continue

            if isinstance(v, (_dt.datetime, _dt.date)):
                v = v.strftime("%Y-%m-%d")

            if field.startswith("_"):
                aux[field] = str(v).strip()
                continue

            if field == "semestre":
                v = str(v)[:10]
            elif field == "catalogo":
                v = str(v)[:64]
            elif field == "asignatura":
                v = str(v)[:255]
            elif field == "creditos":
                try:
                    v = int(float(v))
                except (TypeError, ValueError):
                    continue
            elif field in ("hrs_semanal", "hrs_semestre"):
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    continue
            elif field in ("programa_codigo", "modalidad", "atrib_curso",
                           "cohorte", "grado", "campus"):
                v = str(v)[:50]
            kw[field] = v

        # Filtrar filas inactivas del formato B (Estado 'I' = inactivo)
        if aux.get("_estado") and aux["_estado"].upper().startswith("I"):
            continue

        # Si no hay columna 'Semestre' pero sí 'Descr', extraer el semestre
        if "semestre" not in kw and aux.get("_descr"):
            sem = _parse_semestre_desde_descr(aux["_descr"])
            if sem:
                kw["semestre"] = sem[:10]

        # Los planes sin catálogo o asignatura se descartan
        if "catalogo" not in kw or "asignatura" not in kw:
            continue

        # Semestre: si falta, usar "N/D" para no descartar la fila
        if "semestre" not in kw:
            kw["semestre"] = "N/D"

        if "programa_codigo" not in kw:
            kw["programa_codigo"] = code_default

        rows.append(kw)

    return rows
