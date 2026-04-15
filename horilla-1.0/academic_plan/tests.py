"""
Tests del importador de Plan de Estudios.
Cubre los 2 formatos reales que entrega Elysa.
"""
import io
from django.test import TestCase
from openpyxl import Workbook

from academic_plan.import_helpers import (
    parse_plan, _find_header_row, _parse_semestre_desde_descr,
)


def _buf(wb):
    b = io.BytesIO()
    wb.save(b)
    b.seek(0)
    return b.read()


class ParsePlanFormatoATests(TestCase):
    """Formato A: sheet1 con headers en fila 1."""

    def test_parsea_sheet1_con_headers_fila_1(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"
        ws.append(["Campus", "Prog Acad", "Descripción", "Plan Acad",
                   "Semestre", "Catálogo", "Asignatura",
                   " Creditos", " Hrs Semana", " Hrs Semestre"])
        ws.append(["MONTR", "MADNE", "Maestría Admon", "MADNE20212",
                   "I Semestre", "10003MBA", "CONTABILIDAD GERENCIAL",
                   2, 2, 24])
        ws.append(["MONTR", "MADNE", "Maestría Admon", "MADNE20212",
                   "I Semestre", "10013MBA", "Administración",
                   2, 2, 24])

        rows = parse_plan(_buf(wb), "MADNE-plan.xlsx")
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["catalogo"], "10003MBA")
        self.assertEqual(rows[0]["asignatura"], "CONTABILIDAD GERENCIAL")
        self.assertEqual(rows[0]["creditos"], 2)


class ParsePlanFormatoBTests(TestCase):
    """Formato B: sheet1 con 'Plan de Estudio' en R1 y headers en R2."""

    def test_parsea_con_nom_largo_y_max_uni(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"
        ws.append(["Plan de Estudio", " 3", None, None, None, None, None,
                   None, None, None, None, None, None, None, None])
        ws.append(["Institución", "Campus", "Grado", "Prog Acad",
                   "Descripción", "Plan Acad", "Lista Cso", "Descr",
                   "ID Curso", "Nom Largo", "Máx Uni", "Hrs Cso",
                   "Catálogo", "Estado", "Cso Equiv"])
        ws.append(["USINU", "MONTR", "MSTR", "MMEPI", "Maestría Epidemiología",
                   "MMEPI20212", "000006266", "MMEPI20212 I Semestre",
                   "109121", "Demografía y Salud", 2, 24,
                   "10005MMEPI", "A", None])
        ws.append(["USINU", "MONTR", "MSTR", "MMEPI", "Maestría Epidemiología",
                   "MMEPI20212", "000006266", "MMEPI20212 II Semestre",
                   "109118", "Epidemiología I", 3, 36,
                   "10002MMEPI", "A", None])

        rows = parse_plan(_buf(wb), "epidemiologia.xlsx")
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["catalogo"], "10005MMEPI")
        self.assertEqual(rows[0]["asignatura"], "Demografía y Salud")
        self.assertEqual(rows[0]["creditos"], 2)
        self.assertEqual(rows[0]["hrs_semestre"], 24)
        # Semestre extraído de Descr
        self.assertEqual(rows[0]["semestre"], "I")
        self.assertEqual(rows[1]["semestre"], "II")

    def test_descarta_filas_estado_inactivo(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"
        ws.append(["Plan de Estudio", " 2"])
        ws.append(["Institución", "Campus", "Grado", "Prog Acad",
                   "Descripción", "Plan Acad", "Lista Cso", "Descr",
                   "ID Curso", "Nom Largo", "Máx Uni", "Hrs Cso",
                   "Catálogo", "Estado", "Cso Equiv"])
        ws.append(["USINU", "MONTR", "MSTR", "MMEPI", "Descripcion",
                   "MMEPI20212", "000006266", "MMEPI20212 I Semestre",
                   "999", "Asignatura Activa", 2, 24,
                   "CAT001", "A", None])
        ws.append(["USINU", "MONTR", "MSTR", "MMEPI", "Descripcion",
                   "MMEPI20212", "000006266", "MMEPI20212 I Semestre",
                   "998", "Asignatura Inactiva", 2, 24,
                   "CAT002", "I", None])

        rows = parse_plan(_buf(wb), "test.xlsx")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["catalogo"], "CAT001")

    def test_archivo_vacio_retorna_lista_vacia(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"
        ws.append(["Plan de Estudio", " 0"])
        ws.append(["Institución", "Campus", "Grado", "Prog Acad",
                   "Descripción", "Plan Acad", "Lista Cso", "Descr",
                   "ID Curso", "Nom Largo", "Máx Uni", "Hrs Cso",
                   "Catálogo", "Estado", "Cso Equiv"])
        rows = parse_plan(_buf(wb), "vacio.xlsx")
        self.assertEqual(rows, [])


class SemestreRegexTests(TestCase):
    def test_romanos(self):
        self.assertEqual(_parse_semestre_desde_descr("MMEPI20212 I Semestre"), "I")
        self.assertEqual(_parse_semestre_desde_descr("MMEPI20212 II Semestre"), "II")
        self.assertEqual(_parse_semestre_desde_descr("MMEPI20212 III Semestre"), "III")
        self.assertEqual(_parse_semestre_desde_descr("MMEPI20212 IV Semestre"), "IV")

    def test_sin_semestre(self):
        self.assertEqual(_parse_semestre_desde_descr(""), "")
        self.assertEqual(_parse_semestre_desde_descr("texto sin nada"), "")
