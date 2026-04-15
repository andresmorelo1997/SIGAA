"""
academic_load.import_helpers
============================
Lógica de importación de archivos Excel del SIGAA — equivalente Python al
`src/lib/import-helpers.ts` del SIGAA Next.js.

Maneja tres tipos de archivos:
  - US_PROG_CLASES_*.xlsx  → carga académica (clases)
  - LC_PROGRAMACION_*.xlsx → programación con docente
  - US_DATOS_DOCENTES_*.xlsx → datos docentes

Cada archivo Elysa trae un encabezado con metadata (ciclo_lectivo, grado,
campus). El parser detecta el tipo, encuentra la fila de headers, parsea
las filas y permite filtrar por campus seleccionados.
"""
from __future__ import annotations

import io
from typing import Iterable, Optional

from openpyxl import load_workbook


def detect_file_type(filename: str) -> str:
    name = (filename or "").upper()
    if "US_PROG_CLASES" in name or "CARGA_ACADEMICA" in name or "CARGA ACADÉMICA" in name or "CARGA ACADEMICA" in name:
        return "US_PROG"
    if "LC_PROG" in name:
        return "LC_PROG"
    if "US_DATOS_DOCENTES" in name or "US_DATOS" in name:
        return "US_DATOS"
    if "DOCENTES_IES" in name:
        return "DOCENTES_IES"
    if "CORTE" in name and "PRENOMINA" in name:
        return "CORTE_PRENOMINA"
    return "UNKNOWN"


def pick_best_sheet(wb):
    """Elige la hoja con más filas — los Excel reales del usuario tienen
    metadata en Hoja1 y los datos en Hoja2."""
    best = wb.active
    best_size = best.max_row * best.max_column
    for ws in wb.worksheets:
        size = (ws.max_row or 0) * (ws.max_column or 0)
        if size > best_size:
            best = ws
            best_size = size
    return best


def find_header_row(sheet, max_search: int = 15) -> int:
    """Localiza la fila de headers (la que contiene 'Nº Clase', 'Catálogo',
    'Instructor', 'Campus', etc.)."""
    KEYS = ("nº clase", "no clase", "catálogo", "catalogo", "instructor",
            "campus", "ccl lvo", "ciclo lectivo", "asignatura")
    for i, row in enumerate(sheet.iter_rows(max_row=max_search, values_only=True)):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " | ".join(cells)
        if sum(1 for k in KEYS if k in joined) >= 2:
            return i
    return 0


def normalize_campus(value) -> str:
    if not value:
        return ""
    s = str(value).strip().upper()
    # Map common variants
    if s in ("MONTERIA", "MONTERÍA"):
        return "MONTR"
    if s in ("CARTAGENA",):
        return "CARTG"
    if s in ("BOGOTA", "BOGOTÁ"):
        return "BOGT"
    return s[:10]


def preview(file_bytes: bytes, filename: str) -> dict:
    """Lee el archivo y devuelve metadata sin insertar nada."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = pick_best_sheet(wb)

    file_type = detect_file_type(filename)
    header_row = find_header_row(sheet)
    headers = []
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True):
        headers = [str(c).strip() if c is not None else "" for c in row]
        break

    def find_col(*candidates) -> int:
        for i, h in enumerate(headers):
            for cand in candidates:
                if h.strip().lower() == cand.lower():
                    return i
        return -1

    campus_col = find_col("Campus", "Sede")
    grado_col = find_col("Grado", "Nivel")
    ciclo_col = find_col("Ciclo Lectivo", "Ciclo", "Periodo")

    campus_set = set()
    grado_set = set()
    ciclo = ""
    total = 0

    for row in sheet.iter_rows(min_row=header_row + 2, values_only=True):
        if not any(c not in (None, "") for c in row):
            continue
        total += 1
        if campus_col >= 0:
            v = row[campus_col]
            if v not in (None, ""):
                campus_set.add(normalize_campus(v))
        if grado_col >= 0:
            v = row[grado_col]
            if v not in (None, ""):
                grado_set.add(str(v).strip()[:10])
        if not ciclo and ciclo_col >= 0:
            v = row[ciclo_col]
            if v not in (None, ""):
                ciclo = str(v).strip()[:10]

    wb.close()

    return {
        "file_type": file_type,
        "total_rows": total,
        "campus": sorted(campus_set),
        "grados": sorted(grado_set),
        "ciclo_lectivo": ciclo,
        "filename": filename,
    }


# Mapeo Excel → modelo (US_PROG_CLASES) — incluye variantes Elysa
COLMAP_US_PROG = {
    "Nº Clase": "num_clase",
    "No Clase": "num_clase",
    "ID Curso": "id_curso",
    "Crd": "creditos",
    "Creditos": "creditos",
    "Sección": "seccion",
    "Estado Clase": "estado_clase",
    "Catálogo": "catalogo",
    "Asignaturas": "descripcion",
    "Descripción": "descripcion",
    "Componente": "componente",
    "Cpte": "componente",
    "Campus": "campus",
    "Grado": "grado",
    "Ciclo Lectivo": "ciclo_lectivo",
    "Ccl Lvo": "ciclo_lectivo",
    "Grupo Académico": "grupo_academico",
    "Organización Académica": "org_academica",
    "Org. Academica": "desc_org_academica",
    "F Inicial": "fecha_inicial",
    "Fecha Final": "fecha_final",
    "Semanas": "semanas",
    "Hora Inicio": "hora_inicio",
    "Hora Fin": "hora_fin",
    "Jornada": "jornada",
    "Lunes": "lunes",
    "Martes": "martes",
    "Miércoles": "miercoles",
    "Jueves": "jueves",
    "Viernes": "viernes",
    "Sábado": "sabado",
    "Domingo": "domingo",
    "Cupos": "capacidad_inscripcion",
    "Capcidad Inscripción": "capacidad_inscripcion",
    "Total Inscripciones": "total_inscripciones",
    "Nombre Instructor": "nombre_instructor",
    "Instructor ID": "instructor_id_raw",
    "ID Instructor": "instructor_id_raw",
    "Instructor": "instructor_id_raw",  # ← variante de Elysa real
    "Hrs Semanal": "hrs_semanal",
    "Hrs Semestre": "hrs_semestre",
    "Horas c/Profesor": "hrs_semanal",   # ← variante real
    "Horas Carga Trabj": "hrs_semestre", # ← variante real
}


def parse_rows(file_bytes: bytes, allowed_campus: Optional[Iterable[str]] = None) -> list[dict]:
    """Parsea US_PROG_CLASES y retorna lista de dicts con kwargs para el modelo.

    Si `allowed_campus` se pasa, solo retorna filas cuyo campus esté en el set.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = pick_best_sheet(wb)
    header_row = find_header_row(sheet)
    headers = []
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True):
        headers = [str(c).strip() if c is not None else "" for c in row]
        break

    allowed = set(allowed_campus) if allowed_campus is not None else None
    rows_out: list[dict] = []
    import datetime as _dt

    for row in sheet.iter_rows(min_row=header_row + 2, values_only=True):
        if not any(c not in (None, "") for c in row):
            continue
        kwargs: dict = {}
        for i, header in enumerate(headers):
            field = COLMAP_US_PROG.get(header.strip())
            if not field or i >= len(row):
                continue
            v = row[i]
            if v in (None, ""):
                continue
            if field == "campus":
                v = normalize_campus(v)
            # Fechas/horas → string corto
            if isinstance(v, (_dt.datetime, _dt.date)):
                v = v.strftime("%Y-%m-%d")
            elif isinstance(v, _dt.time):
                v = v.strftime("%H:%M")
            # Truncar strings a límites razonables
            if field in ("estado_clase",):
                v = str(v)[:20]
            elif field in ("componente",):
                v = str(v)[:16]
            elif field in ("catalogo",):
                v = str(v)[:64]
            elif field in ("descripcion", "nombre_instructor", "desc_org_academica"):
                v = str(v)[:255]
            elif field in ("campus", "grado"):
                v = str(v)[:10]
            elif field in ("ciclo_lectivo",):
                v = str(v)[:10]
            elif field in ("grupo_academico", "org_academica"):
                v = str(v)[:64]
            elif field in ("hora_inicio", "hora_fin"):
                v = str(v)[:10]
            elif field in ("jornada",):
                v = str(v)[:20]
            elif field in ("lunes", "martes", "miercoles", "jueves",
                           "viernes", "sabado", "domingo"):
                v = str(v)[:2]
            elif field in ("fecha_inicial", "fecha_final"):
                v = str(v)[:20]
            elif field in ("instructor_id_raw",):
                v = str(v).strip()
            elif field in ("num_clase", "id_curso", "creditos", "seccion",
                           "semanas", "capacidad_inscripcion", "total_inscripciones"):
                try: v = int(float(v))
                except (TypeError, ValueError): continue
            elif field in ("hrs_semanal", "hrs_semestre"):
                try: v = float(v)
                except (TypeError, ValueError): continue
            kwargs[field] = v

        if allowed and kwargs.get("campus") not in allowed:
            continue

        # Padding instructor_id a 10 dígitos
        ins = kwargs.get("instructor_id_raw")
        if ins is not None:
            s = str(ins).strip()
            if s.isdigit() and len(s) < 10:
                kwargs["instructor_id_raw"] = s.zfill(10)
            else:
                kwargs["instructor_id_raw"] = s

        rows_out.append(kwargs)
    wb.close()
    return rows_out
