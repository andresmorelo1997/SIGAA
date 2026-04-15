"""
academic_load.views
===================
Vistas del módulo Carga Académica:
  - list view con filtros multi-select estilo Excel
  - import wizard (upload → preview → confirm)
  - history view (bitácora con descarga)
  - download del archivo original
"""
from __future__ import annotations

import io
import os
import tempfile
import uuid

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.views.decorators.http import require_http_methods

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from base.models import Department
from employee.models import Employee

from .forms import FilterForm, ImportConfirmForm, ImportUploadForm
from .import_helpers import detect_file_type, parse_rows, preview
from .models import (
    CAMPUS_CHOICES,
    GRADO_CHOICES,
    CargaAcademica,
    ImportHistory,
)


# ---------------------------------------------------------------- LIST
@login_required
def carga_list(request):
    qs = CargaAcademica.objects.select_related("docente", "programa").all()

    form = FilterForm(request.GET or None)
    if form.is_valid():
        d = form.cleaned_data
        if d.get("search"):
            term = d["search"]
            qs = qs.filter(
                Q(descripcion__icontains=term)
                | Q(catalogo__icontains=term)
                | Q(nombre_instructor__icontains=term)
                | Q(instructor_id_raw__icontains=term)
            )
        if d.get("campus"):
            qs = qs.filter(campus__in=d["campus"])
        if d.get("grado"):
            qs = qs.filter(grado__in=d["grado"])
        if d.get("estado"):
            qs = qs.filter(estado_clase__in=d["estado"])
        if d.get("ciclo_lectivo"):
            qs = qs.filter(ciclo_lectivo=d["ciclo_lectivo"])
        if d.get("rango_horas"):
            try:
                lo, hi = [int(x) for x in d["rango_horas"].split("-")]
                qs = qs.filter(hrs_semanal__gte=lo, hrs_semanal__lte=hi)
            except (ValueError, AttributeError):
                pass

    paginator = Paginator(qs, 50)
    page = paginator.get_page(request.GET.get("page") or 1)
    return render(
        request,
        "academic_load/carga_list.html",
        {
            "page": page,
            "filter_form": form,
            "total": paginator.count,
            "campus_choices": CAMPUS_CHOICES,
            "grado_choices": GRADO_CHOICES,
        },
    )


# ---------------------------------------------------------------- EXPORT
@login_required
def carga_export(request):
    """Exporta la carga académica filtrada a Excel (mismos filtros que /academic/)."""
    qs = CargaAcademica.objects.select_related("docente", "programa").all()
    form = FilterForm(request.GET or None)
    if form.is_valid():
        d = form.cleaned_data
        if d.get("search"):
            t = d["search"]
            qs = qs.filter(
                Q(descripcion__icontains=t) | Q(catalogo__icontains=t)
                | Q(nombre_instructor__icontains=t) | Q(instructor_id_raw__icontains=t)
            )
        if d.get("campus"): qs = qs.filter(campus__in=d["campus"])
        if d.get("grado"): qs = qs.filter(grado__in=d["grado"])
        if d.get("estado"): qs = qs.filter(estado_clase__in=d["estado"])
        if d.get("ciclo_lectivo"): qs = qs.filter(ciclo_lectivo=d["ciclo_lectivo"])
        if d.get("rango_horas"):
            try:
                lo, hi = [int(x) for x in d["rango_horas"].split("-")]
                qs = qs.filter(hrs_semanal__gte=lo, hrs_semanal__lte=hi)
            except (ValueError, AttributeError):
                pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Carga Académica"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="A6192E")
    headers = ["N°", "N° Clase", "Catálogo", "Asignatura", "Ciclo", "Campus",
               "Grado", "Cédula docente", "Docente", "Hrs/sem", "Hrs/sem.re",
               "Estado"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for i, c in enumerate(qs.iterator(), start=1):
        ws.append([
            i,
            c.num_clase or "",
            c.catalogo or "",
            c.descripcion or "",
            c.ciclo_lectivo or "",
            c.campus or "",
            c.grado or "",
            c.instructor_id_raw or "",
            c.nombre_instructor or "",
            c.hrs_semanal or 0,
            c.hrs_semestre or 0,
            c.estado_clase or "",
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="carga_academica.xlsx"'
    return resp


# ---------------------------------------------------------------- IMPORT
PENDING_DIR = os.path.join(tempfile.gettempdir(), "academic_load_pending")
os.makedirs(PENDING_DIR, exist_ok=True)


@login_required
@require_http_methods(["GET", "POST"])
def import_upload(request):
    if request.method == "POST":
        form = ImportUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data["archivo"]
            data = f.read()
            token = uuid.uuid4().hex
            path = os.path.join(PENDING_DIR, f"{token}.xlsx")
            with open(path, "wb") as out:
                out.write(data)
            request.session["pending_import"] = {
                "filename": f.name,
                "token": token,
            }
            return HttpResponseRedirect(reverse("academic-load-import-preview"))
    else:
        form = ImportUploadForm()
    return render(request, "academic_load/import_upload.html", {"form": form})


@login_required
def import_preview(request):
    pending = request.session.get("pending_import")
    if not pending:
        messages.warning(request, "Primero suba un archivo Excel.")
        return HttpResponseRedirect(reverse("academic-load-import"))

    path = os.path.join(PENDING_DIR, f"{pending['token']}.xlsx")
    if not os.path.exists(path):
        messages.warning(request, "El archivo se perdió. Vuelva a subirlo.")
        return HttpResponseRedirect(reverse("academic-load-import"))

    with open(path, "rb") as fh:
        file_bytes = fh.read()
    info = preview(file_bytes, pending["filename"])

    # Default: deselect CARTG (Cartagena)
    default_campus = [c for c in info["campus"] if c != "CARTG"]

    if request.method == "POST":
        chosen = request.POST.getlist("campus_ids")
        if not chosen:
            messages.error(request, "Debe seleccionar al menos un campus para importar.")
        else:
            result = _do_import(request, file_bytes, pending["filename"], info, chosen)
            # Cleanup tmp
            try: os.remove(path)
            except: pass
            return result

    return render(
        request,
        "academic_load/import_preview.html",
        {
            "info": info,
            "default_campus": default_campus,
        },
    )


def _do_import(request, file_bytes: bytes, filename: str, info: dict, campus_chosen: list[str]):
    """Importa el Excel: crea docentes/empleados que falten + filas de carga."""
    from django.contrib.auth.models import User

    file_type = info["file_type"]
    history = ImportHistory.objects.create(
        filename=filename, file_type=file_type, file_size=len(file_bytes),
        file_blob=file_bytes, ciclo_lectivo=info["ciclo_lectivo"],
        campus=",".join(campus_chosen), status="processing",
    )

    docentes_creados = 0
    inserted = 0
    skipped = 0
    try:
        rows = parse_rows(file_bytes, allowed_campus=set(campus_chosen))

        # Build employee lookup by badge_id
        emp_by_badge = {
            e.badge_id: e
            for e in Employee.objects.exclude(badge_id__isnull=True).exclude(badge_id="")
        }

        objs = []
        for r in rows:
            ins = r.get("instructor_id_raw")
            nombre = (r.get("nombre_instructor") or "").strip()

            # Si el archivo trae un docente nuevo (instructor_id + nombre),
            # crear automáticamente Employee + User para vincular la clase.
            if ins and ins not in emp_by_badge and nombre:
                # Parsear nombre "Apellido,Nombre" o "Nombre Apellido"
                if "," in nombre:
                    apellido, primer_nombre = [x.strip() for x in nombre.split(",", 1)]
                else:
                    parts = nombre.split()
                    primer_nombre = parts[0] if parts else "Docente"
                    apellido = " ".join(parts[1:]) if len(parts) > 1 else primer_nombre

                username = f"docente_{ins}"
                user, _ = User.objects.get_or_create(
                    username=username,
                    defaults={
                        "email": f"{ins}@unisinu.edu.co",
                        "first_name": primer_nombre[:30],
                        "last_name": apellido[:30],
                        "is_active": True,
                    },
                )
                if not user.password:
                    user.set_password(f"docente{ins[-4:]}")
                    user.save()

                try:
                    emp = Employee.objects.create(
                        employee_user_id=user,
                        employee_first_name=primer_nombre[:200],
                        employee_last_name=apellido[:200],
                        email=f"{ins}@unisinu.edu.co",
                        phone="0",
                        country="Colombia",
                        badge_id=ins[:50],
                    )
                    emp_by_badge[ins] = emp
                    docentes_creados += 1
                except Exception:
                    pass

            if ins and ins in emp_by_badge:
                r["docente"] = emp_by_badge[ins]
            r["import_id"] = history
            objs.append(CargaAcademica(**r))

        CargaAcademica.objects.bulk_create(objs, batch_size=500)
        inserted = len(objs)
        skipped = info["total_rows"] - inserted

        history.records_inserted = inserted
        history.records_skipped = skipped
        history.status = "completed"
        history.save()

        request.session.pop("pending_import", None)

        total_archivo = info["total_rows"]
        pct = round(100 * inserted / total_archivo, 1) if total_archivo else 0
        msg = (
            f"Importación completa: {inserted} de {total_archivo} clases ({pct}%) insertadas"
            + (f" · {docentes_creados} docentes nuevos creados" if docentes_creados else "")
            + (f" · {skipped} omitidas por filtro de campus" if skipped else "")
            + "."
        )
        messages.success(request, msg)
        return HttpResponseRedirect(reverse("academic-load-list"))
    except Exception as e:
        history.status = "error"
        history.error_message = str(e)
        history.save()
        messages.error(request, f"Error al importar: {e}")
        return HttpResponseRedirect(reverse("academic-load-import"))


@login_required
def history_list(request):
    history = ImportHistory.objects.all()[:200]
    return render(request, "academic_load/history.html", {"history": history})


# ---------------------------------------------------------------- BUSQUEDA GLOBAL
@login_required
def busqueda_global(request):
    """Busca docentes por nombre o cédula y redirige o lista resultados."""
    from employee.models import Employee
    q = (request.GET.get("q") or "").strip()
    if not q:
        return HttpResponseRedirect(reverse("academic-dashboard"))

    candidatos = Employee.objects.filter(
        Q(employee_first_name__icontains=q)
        | Q(employee_last_name__icontains=q)
        | Q(badge_id__icontains=q)
        | Q(email__icontains=q)
    )[:50]
    # Si hay exactamente uno, redirige al detail
    count = candidatos.count()
    if count == 1:
        return HttpResponseRedirect(reverse("employee-view-individual", args=[candidatos[0].id]))

    return render(request, "academic_load/busqueda.html", {
        "q": q,
        "resultados": candidatos,
        "total": count,
    })


# ---------------------------------------------------------------- DASHBOARD SIGAA
@login_required
def dashboard_sigaa(request):
    """Dashboard integrador: métricas cruzadas del sistema académico."""
    from django.db.models import Sum, Count
    from academic_plan.models import PlanEstudio
    try:
        from employee.models import Employee
    except Exception:
        Employee = None

    ciclos = sorted(set(
        CargaAcademica.objects.exclude(ciclo_lectivo__isnull=True)
        .exclude(ciclo_lectivo="").values_list("ciclo_lectivo", flat=True)
    ), reverse=True)
    ciclo = request.GET.get("ciclo", ciclos[0] if ciclos else "")

    carga_qs = CargaAcademica.objects.all()
    if ciclo:
        carga_qs = carga_qs.filter(ciclo_lectivo=ciclo)

    # KPI
    kpis = {
        "total_docentes": (Employee.objects.count() if Employee else 0),
        "total_clases": carga_qs.count(),
        "total_asignaturas_plan": PlanEstudio.objects.count(),
        "total_programas": PlanEstudio.objects.values("programa_codigo").distinct().count(),
        "hrs_totales": round(carga_qs.aggregate(s=Sum("hrs_semestre"))["s"] or 0, 1),
        "docentes_con_carga": carga_qs.exclude(docente__isnull=True).values("docente").distinct().count(),
    }

    # Cobertura plan vs carga
    plan_cats = set(PlanEstudio.objects.values_list("catalogo", flat=True))
    carga_cats = set(carga_qs.values_list("catalogo", flat=True))
    cobertura_total = len(plan_cats & carga_cats)
    kpis["cobertura_asignaturas"] = cobertura_total
    kpis["cobertura_pct"] = round(100 * cobertura_total / len(plan_cats), 1) if plan_cats else 0

    # Top docentes por horas
    top_docentes = list(
        carga_qs.exclude(docente__isnull=True)
        .values("docente__id", "docente__employee_first_name", "docente__employee_last_name")
        .annotate(hrs=Sum("hrs_semestre"), clases=Count("id"))
        .order_by("-hrs")[:10]
    )

    # Distribución por campus
    por_campus = list(
        carga_qs.exclude(campus__isnull=True).exclude(campus="")
        .values("campus").annotate(n=Count("id"), hrs=Sum("hrs_semestre"))
        .order_by("-n")
    )
    max_campus = max((c["n"] for c in por_campus), default=1)
    for c in por_campus:
        c["pct"] = round(100 * c["n"] / max_campus, 1)

    # Distribución por grado
    por_grado = list(
        carga_qs.exclude(grado__isnull=True).exclude(grado="")
        .values("grado").annotate(n=Count("id"), hrs=Sum("hrs_semestre"))
        .order_by("-n")
    )
    max_grado = max((g["n"] for g in por_grado), default=1)
    for g in por_grado:
        g["pct"] = round(100 * g["n"] / max_grado, 1)

    # Max horas de los top docentes para la barra
    max_hrs_top = max((d["hrs"] or 0 for d in top_docentes), default=1)
    for d in top_docentes:
        d["pct"] = round(100 * (d["hrs"] or 0) / max_hrs_top, 1)

    # Programas críticos (cobertura < 50%)
    from collections import defaultdict
    plan_por_prog = defaultdict(set)
    for cat, prog in PlanEstudio.objects.values_list("catalogo", "programa_codigo"):
        plan_por_prog[prog].add(cat)
    programas_criticos = []
    for prog, cats in plan_por_prog.items():
        if not cats:
            continue
        cubiertos = len(cats & carga_cats)
        pct = 100 * cubiertos / len(cats)
        if pct < 50:
            programas_criticos.append({
                "programa": prog,
                "total": len(cats),
                "cubiertos": cubiertos,
                "faltantes": len(cats) - cubiertos,
                "pct": round(pct, 1),
            })
    programas_criticos.sort(key=lambda x: x["pct"])

    return render(request, "academic_load/dashboard_sigaa.html", {
        "kpis": kpis,
        "ciclos": ciclos,
        "ciclo": ciclo,
        "top_docentes": top_docentes,
        "por_campus": por_campus,
        "por_grado": por_grado,
        "programas_criticos": programas_criticos[:5],
    })


# ---------------------------------------------------------------- EMPLOYEE TAB
@login_required
def academic_load_employee_tab(request, emp_id: int):
    """Renderiza la pestaña 'Carga Académica' dentro del detalle de un docente.

    Busca clases por FK `docente` o por `instructor_id_raw` (cédula padded
    a 10 dígitos). Devuelve todas las clases históricas del docente
    agrupadas por ciclo lectivo.
    """
    from django.db.models import Sum, Count
    emp = get_object_or_404(Employee, pk=emp_id)

    # Cruce robusto: por FK o por instructor_id_raw (cédula)
    cedula_raw = (emp.badge_id or "").strip()
    cedula_padded = cedula_raw.zfill(10) if cedula_raw.isdigit() else cedula_raw
    clases = CargaAcademica.objects.filter(
        Q(docente_id=emp.id)
        | Q(instructor_id_raw=cedula_raw)
        | Q(instructor_id_raw=cedula_padded)
    ).order_by("-ciclo_lectivo", "catalogo")

    # Resumen por ciclo
    resumen = (
        clases.values("ciclo_lectivo", "grado", "campus")
        .annotate(
            n_clases=Count("id"),
            hrs_semanal=Sum("hrs_semanal"),
            hrs_semestre=Sum("hrs_semestre"),
        )
        .order_by("-ciclo_lectivo", "grado")
    )

    totales = clases.aggregate(
        n=Count("id"),
        hs=Sum("hrs_semanal"),
        hsr=Sum("hrs_semestre"),
    )

    return render(request, "academic_load/tab_carga_docente.html", {
        "employee": emp,
        "clases": clases[:300],
        "resumen": resumen,
        "totales": totales,
    })


@login_required
def history_download(request, pk: int):
    h = get_object_or_404(ImportHistory, pk=pk)
    if not h.file_blob:
        messages.warning(request, "Este registro no tiene archivo guardado.")
        return HttpResponseRedirect(reverse("academic-load-history"))
    resp = HttpResponse(
        bytes(h.file_blob),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{h.filename}"'
    return resp
