"""
academic_payroll.views
======================
Vistas del módulo Prenómina Docente.
"""
import io
from datetime import datetime, date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.urls import reverse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from django.db.models import Q

from academic_load.models import CargaAcademica, GRADO_CHOICES, CAMPUS_CHOICES
from employee.models import Employee
from .forms import CorteForm
from .models import Corte, PrenominaDocente


# ──────────────────────── helpers ────────────────────────
def _grados_for_tipo(tipo: str) -> list[str]:
    if tipo.upper() == "PREGRADO":
        return ["PREG", "PRE2", "TCN"]
    return ["POSG", "MSTR", "DOCT", "ESP"]


def _parse_fecha(s):
    """Parse dd/mm/yyyy, yyyy-mm-dd, dd-mm-yyyy → date | None."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def _calc_horas_por_corte(hrs_semestre, fecha_ini, fecha_fin, cortes_list):
    """
    Distribuye hrs_semestre entre los cortes, proporcional al overlap de días
    entre la duración de la clase (fecha_ini→fecha_fin) y cada corte.

    Si falta info, cae al reparto uniforme.
    Retorna lista de enteros (redondeados) alineada con cortes_list.
    """
    n = len(cortes_list)
    if n == 0:
        return []
    hrs = float(hrs_semestre or 0)
    if hrs <= 0:
        return [0] * n

    fi = _parse_fecha(fecha_ini)
    ff = _parse_fecha(fecha_fin)

    # Fallback uniforme
    if not fi or not ff or ff < fi:
        base = round(hrs / n)
        res = [base] * n
        res[-1] += int(round(hrs - base * n))
        return res

    total_days = (ff - fi).days + 1
    if total_days <= 0:
        return [round(hrs / n)] * n

    raw = []
    for c in cortes_list:
        ci, cf = c.get("fecha_inicio"), c.get("fecha_fin")
        if not ci or not cf:
            raw.append(0.0)
            continue
        start = max(fi, ci)
        end = min(ff, cf)
        if end < start:
            raw.append(0.0)
        else:
            overlap = (end - start).days + 1
            raw.append(hrs * overlap / total_days)

    # Redondear y ajustar residual al último corte con valor > 0
    rounded = [int(round(x)) for x in raw]
    diff = int(round(hrs)) - sum(rounded)
    if diff != 0:
        for i in range(n - 1, -1, -1):
            if rounded[i] > 0:
                rounded[i] += diff
                break
        else:
            rounded[-1] += diff
    return rounded


def _derive_dedicacion(emp):
    """Devuelve 'Tiempo Completo' | 'Medio Tiempo' | 'Catedrático' | None."""
    try:
        ewi = getattr(emp, "employee_work_info", None)
        if ewi and ewi.job_position_id:
            name = (ewi.job_position_id.job_position or "").strip()
            low = name.lower()
            if "tiempo completo" in low:
                return "Tiempo Completo"
            if "medio tiempo" in low:
                return "Medio Tiempo"
            if "catedr" in low or "hora c" in low:
                return "Catedrático"
            return name or None
    except Exception:
        pass
    return None


def _dedicacion_by_hrs(total_hrs_semana):
    """Fallback: clasifica por hrs/semana totales del docente."""
    h = float(total_hrs_semana or 0)
    if h >= 40:
        return "Tiempo Completo"
    if h >= 20:
        return "Medio Tiempo"
    return "Catedrático"


DEDICACION_ORDER = {"Catedrático": 0, "Medio Tiempo": 1, "Tiempo Completo": 2}


def _calcular_consolidado(periodo: str, tipo: str, ciclo: str = ""):
    """
    Calcula consolidado on-demand desde CargaAcademica.
    Devuelve (rows[], cortes_info[]).
    """
    grados = _grados_for_tipo(tipo)

    cortes = list(Corte.objects.filter(periodo=periodo, grado__in=grados)
                  .order_by("num_corte").values("num_corte", "fecha_inicio",
                                                "fecha_fin", "grado", "emitido"))

    # Agrupa por num_corte (algunos grados comparten corte)
    cortes_info = {}
    for c in cortes:
        n = c["num_corte"]
        if n not in cortes_info:
            cortes_info[n] = {
                "num": n,
                "fecha_inicio": c["fecha_inicio"],
                "fecha_fin": c["fecha_fin"],
                "emitido": c["emitido"],
            }

    cortes_list = sorted(cortes_info.values(), key=lambda x: x["num"])
    num_cortes = len(cortes_list)

    qs = CargaAcademica.objects.filter(grado__in=grados).select_related("docente")
    if ciclo:
        qs = qs.filter(ciclo_lectivo=ciclo)

    # Agrupa por docente
    consol = {}
    for c in qs:
        if not c.docente_id:
            continue
        key = c.docente_id
        if key not in consol:
            consol[key] = {
                "docente": c.docente,
                "campus": c.campus or "",
                "programa": c.programa.department if c.programa else "",
                "hrs_semana": 0,
                "hrs_semestre": 0,
                "cortes": [0] * 6,
                "clases": [],
            }
        d = consol[key]
        d["hrs_semana"] += float(c.hrs_semanal or 0)
        d["hrs_semestre"] += float(c.hrs_semestre or 0)
        # Reparto sencillo de horas por corte (proporcional al número de cortes)
        if num_cortes > 0:
            por_corte = (c.hrs_semestre or 0) / num_cortes
            for i in range(num_cortes):
                d["cortes"][i] += por_corte

        d["clases"].append({
            "num_clase": c.num_clase, "catalogo": c.catalogo,
            "asignatura": c.descripcion, "componente": c.componente,
            "hrs_semana": c.hrs_semanal or 0, "hrs_semestre": c.hrs_semestre or 0,
        })

    rows = []
    for d in consol.values():
        hrs_prenom = round(sum(d["cortes"][:num_cortes]), 2)
        rows.append({
            "docente_id": d["docente"].id,
            "instructor_id": d["docente"].badge_id,
            "nombre": f"{d['docente'].employee_first_name} {d['docente'].employee_last_name}",
            "campus": d["campus"], "programa": d["programa"],
            "hrs_semana": round(d["hrs_semana"], 2),
            "hrs_semestre": round(d["hrs_semestre"], 2),
            "cortes": [round(x, 2) for x in d["cortes"][:num_cortes]],
            "hrs_prenomina": hrs_prenom,
            "saldo": round(d["hrs_semestre"] - hrs_prenom, 2),
        })
    rows.sort(key=lambda x: x["nombre"])

    return rows, cortes_list


# ──────────────────────── views ────────────────────────
@login_required
def consolidado(request):
    periodo = request.GET.get("periodo", "2026-1")
    tipo = request.GET.get("tipo", "PREGRADO")
    ciclo = request.GET.get("ciclo", "")
    corte_hasta = request.GET.get("corte")  # Filtrar solo hasta el corte N
    q = (request.GET.get("q") or "").strip().lower()

    # opciones disponibles
    periodos = sorted(set(Corte.objects.values_list("periodo", flat=True))) or ["2026-1"]
    ciclos = sorted(set(
        CargaAcademica.objects.exclude(ciclo_lectivo__isnull=True)
        .exclude(ciclo_lectivo="").values_list("ciclo_lectivo", flat=True)
    ))

    rows, cortes_info = _calcular_consolidado(periodo, tipo, ciclo)

    # Limitar cortes a mostrar: si se pide "hasta corte N", solo C1..CN
    if corte_hasta:
        try:
            n = int(corte_hasta)
            cortes_info = [c for c in cortes_info if c["num"] <= n]
            # Recalcular hrs_prenomina acumulado hasta el corte N
            for r in rows:
                r["cortes"] = r["cortes"][:n]
                r["hrs_prenomina"] = round(sum(r["cortes"]), 2)
                r["saldo"] = round(r["hrs_semestre"] - r["hrs_prenomina"], 2)
        except ValueError:
            pass

    # Búsqueda por nombre o cédula
    if q:
        rows = [r for r in rows if q in (r["nombre"] or "").lower() or q in (r["instructor_id"] or "").lower()]

    return render(request, "academic_payroll/consolidado.html", {
        "rows": rows, "cortes_info": cortes_info,
        "periodo": periodo, "tipo": tipo, "ciclo": ciclo,
        "corte_hasta": corte_hasta, "q": q,
        "periodos": periodos, "ciclos": ciclos,
        "total": len(rows),
    })


@login_required
def detalle(request):
    """
    Detalle de Prenómina — formato UniSinú (una fila por clase,
    agrupado por Dedicación Académica → Docente con subtotales).
    """
    periodo = request.GET.get("periodo", "2026-1")
    corte_sel = request.GET.get("corte") or ""     # 1..4 o "" para mostrar todos
    dedicacion_sel = request.GET.get("dedicacion", "")
    campus_sel = request.GET.get("campus", "")
    q = (request.GET.get("q") or "").strip().lower()

    # Cortes únicos por num_corte (tomamos las fechas del primero que aparezca)
    cortes_qs = Corte.objects.filter(periodo=periodo).order_by("num_corte")
    cortes_map = {}
    for c in cortes_qs:
        if c.num_corte not in cortes_map:
            cortes_map[c.num_corte] = {
                "num": c.num_corte,
                "fecha_inicio": c.fecha_inicio,
                "fecha_fin": c.fecha_fin,
                "emitido": c.emitido,
            }
    cortes_list = sorted(cortes_map.values(), key=lambda x: x["num"])

    # Rango de fechas global para el subtítulo
    fecha_corte_ini = cortes_list[0]["fecha_inicio"] if cortes_list else None
    fecha_corte_fin = cortes_list[-1]["fecha_fin"] if cortes_list else None

    # Carga académica del periodo (filtramos por ciclo que coincide con periodo si aplica)
    clases = (CargaAcademica.objects
              .select_related("docente", "programa")
              .exclude(docente__isnull=True))
    if campus_sel:
        clases = clases.filter(campus=campus_sel)

    # Agrupar por docente
    by_doc = {}
    for cl in clases:
        d = cl.docente
        if not d:
            continue
        key = d.id
        if key not in by_doc:
            by_doc[key] = {
                "employee": d,
                "documento": (d.badge_id or "").strip(),
                "instructor_id": (d.badge_id or "").strip()[-5:],  # últimos 5
                "nombre": f"{d.employee_first_name} {d.employee_last_name}".strip(),
                "clases": [],
                "hrs_semana": 0.0,
                "hrs_semestre": 0.0,
                "cortes_total": [0] * len(cortes_list),
                "campus": cl.campus or "",
                "programa": cl.programa.department if cl.programa else "",
            }
        grp = by_doc[key]
        hrs_corte = _calc_horas_por_corte(
            cl.hrs_semestre, cl.fecha_inicial, cl.fecha_final, cortes_list
        )
        clase_row = {
            "campus": cl.campus or "",
            "programa": cl.programa.department if cl.programa else "",
            "num_clase": cl.num_clase or "",
            "catalogo": cl.catalogo or "",
            "asignatura": cl.descripcion or "",
            "componente": cl.componente or "",
            "observacion": "" if cl.estado_clase == "Activo" else (cl.estado_clase or ""),
            "hrs_semana": int(round(float(cl.hrs_semanal or 0))),
            "hrs_semestre": int(round(float(cl.hrs_semestre or 0))),
            "fecha_inicial": cl.fecha_inicial or "",
            "fecha_final": cl.fecha_final or "",
            "cortes": hrs_corte,
        }
        grp["clases"].append(clase_row)
        grp["hrs_semana"] += float(cl.hrs_semanal or 0)
        grp["hrs_semestre"] += float(cl.hrs_semestre or 0)
        for i, v in enumerate(hrs_corte):
            grp["cortes_total"][i] += v

    # Derivar dedicación por docente
    for grp in by_doc.values():
        ded = _derive_dedicacion(grp["employee"])
        if not ded:
            ded = _dedicacion_by_hrs(grp["hrs_semana"])
        grp["dedicacion"] = ded

    # Filtro búsqueda y dedicación
    docentes = list(by_doc.values())
    if dedicacion_sel:
        docentes = [d for d in docentes if d["dedicacion"] == dedicacion_sel]
    if q:
        docentes = [d for d in docentes
                    if q in d["nombre"].lower() or q in d["documento"].lower()]

    # Ordenar: por dedicación, luego nombre
    docentes.sort(key=lambda d: (DEDICACION_ORDER.get(d["dedicacion"], 99),
                                  d["nombre"].lower()))

    # Agrupar en secciones por dedicación
    secciones_map = {}
    for d in docentes:
        ded = d["dedicacion"]
        if ded not in secciones_map:
            secciones_map[ded] = {
                "dedicacion": ded,
                "docentes": [],
                "hrs_semana": 0,
                "hrs_semestre": 0,
                "cortes_total": [0] * len(cortes_list),
            }
        secciones_map[ded]["docentes"].append(d)
        secciones_map[ded]["hrs_semana"] += d["hrs_semana"]
        secciones_map[ded]["hrs_semestre"] += d["hrs_semestre"]
        for i, v in enumerate(d["cortes_total"]):
            secciones_map[ded]["cortes_total"][i] += v

    secciones = sorted(secciones_map.values(),
                       key=lambda s: DEDICACION_ORDER.get(s["dedicacion"], 99))

    # Enteros en totales de docente y sección
    for d in docentes:
        d["hrs_semana"] = int(round(d["hrs_semana"]))
        d["hrs_semestre"] = int(round(d["hrs_semestre"]))
    for s in secciones:
        s["hrs_semana"] = int(round(s["hrs_semana"]))
        s["hrs_semestre"] = int(round(s["hrs_semestre"]))

    # Catálogos para filtros
    periodos = sorted(set(Corte.objects.values_list("periodo", flat=True))) or ["2026-1"]
    campuses = [{"code": c[0], "label": c[1]} for c in CAMPUS_CHOICES]

    # Corte Romano para título
    romano = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI"}
    corte_label = romano.get(int(corte_sel), corte_sel) if corte_sel else "—"

    return render(request, "academic_payroll/detalle.html", {
        "secciones": secciones,
        "cortes_list": cortes_list,
        "periodo": periodo,
        "periodos": periodos,
        "corte_sel": corte_sel,
        "corte_label": corte_label,
        "dedicacion_sel": dedicacion_sel,
        "campus_sel": campus_sel,
        "campuses": campuses,
        "q": q,
        "total_docentes": len(docentes),
        "fecha_corte_ini": fecha_corte_ini,
        "fecha_corte_fin": fecha_corte_fin,
        "dedicaciones": ["Catedrático", "Medio Tiempo", "Tiempo Completo"],
    })


@login_required
def cortes_view(request):
    cortes = Corte.objects.all().order_by("periodo", "grado", "num_corte")
    return render(request, "academic_payroll/cortes.html", {"cortes": cortes})


@login_required
def corte_new(request):
    form = CorteForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Corte creado correctamente.")
        return HttpResponseRedirect(reverse("payroll-cortes"))
    return render(request, "academic_payroll/corte_form.html", {"form": form, "is_new": True})


@login_required
def corte_edit(request, pk: int):
    corte = get_object_or_404(Corte, pk=pk)
    form = CorteForm(request.POST or None, instance=corte)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Corte actualizado correctamente.")
        return HttpResponseRedirect(reverse("payroll-cortes"))
    return render(request, "academic_payroll/corte_form.html", {"form": form, "corte": corte})


@login_required
def historicos(request):
    """Muestra snapshots PrenominaDocente emitidos (auditoría)."""
    periodo = request.GET.get("periodo", "")
    tipo = request.GET.get("tipo", "")
    qs = PrenominaDocente.objects.select_related("docente").filter(estado="emitido")
    if periodo:
        qs = qs.filter(periodo=periodo)
    if tipo:
        qs = qs.filter(tipo=tipo)

    periodos = sorted(set(PrenominaDocente.objects.values_list("periodo", flat=True).distinct()))
    tipos = ["PREGRADO", "POSGRADO"]

    return render(request, "academic_payroll/historicos.html", {
        "snapshots": qs[:500],
        "total": qs.count(),
        "periodos": periodos,
        "tipos": tipos,
        "periodo": periodo,
        "tipo": tipo,
    })


@login_required
def corte_emitir(request, pk: int):
    """Emite/congela un corte: crea snapshots PrenominaDocente con las horas
    consolidadas en ese momento. Al re-importar carga académica, estos
    snapshots no cambian."""
    corte = get_object_or_404(Corte, pk=pk)

    if corte.emitido:
        messages.warning(request, f"El corte {corte.num_corte} de {corte.periodo} ya fue emitido.")
        return HttpResponseRedirect(reverse("payroll-cortes"))

    from django.utils import timezone as _tz
    tipo = "POSGRADO" if corte.grado.upper() in ("POSG", "MSTR", "DOCT", "ESP") else "PREGRADO"
    rows, _cortes = _calcular_consolidado(corte.periodo, tipo)

    creados = 0
    for r in rows:
        docente_id = r.get("docente_id")
        if not docente_id:
            continue
        hrs_prenomina = r["hrs_prenomina"]
        try:
            pd, created = PrenominaDocente.objects.update_or_create(
                periodo=corte.periodo,
                tipo=tipo,
                docente_id=docente_id,
                defaults={
                    "hrs_semana": r["hrs_semana"],
                    "hrs_semestre": r["hrs_semestre"],
                    "corte_1": r["cortes"][0] if len(r["cortes"]) > 0 else 0,
                    "corte_2": r["cortes"][1] if len(r["cortes"]) > 1 else 0,
                    "corte_3": r["cortes"][2] if len(r["cortes"]) > 2 else 0,
                    "corte_4": r["cortes"][3] if len(r["cortes"]) > 3 else 0,
                    "corte_5": r["cortes"][4] if len(r["cortes"]) > 4 else 0,
                    "corte_6": r["cortes"][5] if len(r["cortes"]) > 5 else 0,
                    "hrs_prenomina": hrs_prenomina,
                    "saldo": r["saldo"],
                    "estado": "emitido",
                    "num_corte_emitido": corte.num_corte,
                },
            )
            if created:
                creados += 1
        except Exception as e:
            messages.error(request, f"Error en {r['nombre']}: {e}")

    corte.emitido = True
    corte.fecha_emision = _tz.now()
    corte.save(update_fields=["emitido", "fecha_emision"])

    messages.success(
        request,
        f"Corte {corte.num_corte} · {corte.periodo} · {corte.grado} emitido. "
        f"{len(rows)} docentes consolidados ({creados} nuevos snapshots).",
    )
    return HttpResponseRedirect(reverse("payroll-cortes"))


@login_required
def consolidado_export(request):
    """Exporta el consolidado de prenómina (en HORAS) a Excel."""
    periodo = request.GET.get("periodo", "2026-1")
    tipo = request.GET.get("tipo", "PREGRADO")
    ciclo = request.GET.get("ciclo", "")

    rows, cortes_info = _calcular_consolidado(periodo, tipo, ciclo)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Prenomina {periodo}"[:31]

    # Encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="A6192E")  # rojo UniSinú

    headers = ["N°", "Cédula", "Docente", "Campus", "Programa",
               "Hrs/sem", "Hrs/sem.re"]
    for c in cortes_info:
        headers.append(f"Corte {c['num']}")
    headers += ["Hrs prenómina", "Saldo horas"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Filas
    for i, r in enumerate(rows, start=1):
        row = [i, r["instructor_id"], r["nombre"], r["campus"], r["programa"],
               r["hrs_semana"], r["hrs_semestre"]]
        row += r["cortes"]
        row += [r["hrs_prenomina"], r["saldo"]]
        ws.append(row)

    # Auto-ancho básico
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"prenomina_{tipo.lower()}_{periodo}"
    if ciclo:
        filename += f"_{ciclo}"
    filename += ".xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ──────────────────────── EMPLOYEE TAB ────────────────────────
@login_required
def mi_prenomina(request):
    """Shortcut: redirige al employee detail del usuario logueado con tab prenómina.
    Útil para compartir link directo al docente con sus horas."""
    try:
        emp = request.user.employee_get
        if emp:
            return HttpResponseRedirect(
                f"{reverse('employee-view-individual', args=[emp.id])}#academic_payroll_target"
            )
    except Exception:
        pass
    messages.warning(request, "No tiene perfil de docente asociado.")
    return HttpResponseRedirect(reverse("academic-dashboard"))


@login_required
def mi_constancia(request):
    """Shortcut: constancia del docente logueado."""
    try:
        emp = request.user.employee_get
        if emp:
            return HttpResponseRedirect(reverse("payroll-constancia", args=[emp.id]))
    except Exception:
        pass
    messages.warning(request, "No tiene perfil de docente asociado.")
    return HttpResponseRedirect(reverse("academic-dashboard"))


@login_required
def constancia_horas(request, emp_id: int):
    """Constancia imprimible de horas académicas del docente (vista limpia)."""
    emp = get_object_or_404(Employee, pk=emp_id)
    cedula_raw = (emp.badge_id or "").strip()
    cedula_padded = cedula_raw.zfill(10) if cedula_raw.isdigit() else cedula_raw

    clases = CargaAcademica.objects.filter(
        Q(docente_id=emp.id)
        | Q(instructor_id_raw=cedula_raw)
        | Q(instructor_id_raw=cedula_padded)
    ).order_by("-ciclo_lectivo", "catalogo")

    ciclos_qs = (
        clases.exclude(ciclo_lectivo__isnull=True).exclude(ciclo_lectivo="")
        .values("ciclo_lectivo")
        .annotate(n=Count("id"), hs=Sum("hrs_semanal"), hsr=Sum("hrs_semestre"))
        .order_by("-ciclo_lectivo")
    )

    totales = clases.aggregate(n=Count("id"), hs=Sum("hrs_semanal"), hsr=Sum("hrs_semestre"))

    from django.utils import timezone as _tz
    return render(request, "academic_payroll/constancia_horas.html", {
        "employee": emp,
        "ciclos_qs": ciclos_qs,
        "totales": totales,
        "fecha": _tz.now(),
    })


@login_required
def academic_payroll_employee_tab(request, emp_id: int):
    """Renderiza la pestaña 'Prenómina Docente' dentro del detalle del docente.

    Calcula on-demand desde CargaAcademica: hrs/semana, hrs/semestre, cortes
    proporcionales y saldo, TODO en horas (SIGAA no maneja valores monetarios).
    """
    emp = get_object_or_404(Employee, pk=emp_id)

    cedula_raw = (emp.badge_id or "").strip()
    cedula_padded = cedula_raw.zfill(10) if cedula_raw.isdigit() else cedula_raw

    clases = CargaAcademica.objects.filter(
        Q(docente_id=emp.id)
        | Q(instructor_id_raw=cedula_raw)
        | Q(instructor_id_raw=cedula_padded)
    )

    # Filtro opcional por ciclo
    ciclo = request.GET.get("ciclo", "")
    if ciclo:
        clases = clases.filter(ciclo_lectivo=ciclo)

    ciclos_disponibles = sorted(set(
        clases.exclude(ciclo_lectivo__isnull=True).exclude(ciclo_lectivo="")
             .values_list("ciclo_lectivo", flat=True)
    ))

    # Consolidado por ciclo en horas
    filas = []
    for c_ciclo in (ciclos_disponibles if not ciclo else [ciclo]):
        qs = clases.filter(ciclo_lectivo=c_ciclo) if not ciclo else clases
        agg = qs.aggregate(
            hs=Sum("hrs_semanal"),
            hsr=Sum("hrs_semestre"),
            n=Count("id"),
        )
        if not agg["n"]:
            continue
        # Grado dominante para definir cortes
        grado = (qs.values_list("grado", flat=True).first() or "").upper()
        grados_filtro = _grados_for_tipo("POSGRADO" if grado in ("POSG", "MSTR", "DOCT", "ESP") else "PREGRADO")

        cortes = list(
            Corte.objects.filter(
                grado__in=grados_filtro,
            ).order_by("num_corte").values_list("num_corte", flat=True).distinct()
        )
        n_cortes = len(cortes) or 4  # fallback: 4 cortes
        por_corte = round((agg["hsr"] or 0) / n_cortes, 2) if n_cortes else 0
        hrs_cortes = [por_corte] * n_cortes
        hrs_prenomina = round(sum(hrs_cortes), 2)
        saldo = round((agg["hsr"] or 0) - hrs_prenomina, 2)

        filas.append({
            "ciclo": c_ciclo,
            "n_clases": agg["n"],
            "hrs_semana": round(agg["hs"] or 0, 2),
            "hrs_semestre": round(agg["hsr"] or 0, 2),
            "cortes": hrs_cortes,
            "n_cortes": n_cortes,
            "hrs_prenomina": hrs_prenomina,
            "saldo": saldo,
        })

    # Totales global
    totales = clases.aggregate(hs=Sum("hrs_semanal"), hsr=Sum("hrs_semestre"), n=Count("id"))

    return render(request, "academic_payroll/tab_prenomina_docente.html", {
        "employee": emp,
        "filas": filas,
        "totales": totales,
        "ciclos_disponibles": ciclos_disponibles,
        "ciclo_filtro": ciclo,
    })
